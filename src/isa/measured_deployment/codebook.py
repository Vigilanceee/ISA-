"""Post-training measured-IV deployment evaluation for CIFAR-100 ViT fc1 layers.

This script never trains or changes checkpoint parameters.  It calibrates the
actual gate-voltage input of every FFN fc1, selects a per-layer verify voltage,
snaps the continuous EKV branches to the measured cell library, replaces only
the requested fc1 modules, and evaluates an exact piecewise-linear measured-IV
implementation backed by segmented GEMMs.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import os
import random
import time
import warnings
from collections import Counter
from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

os.environ.setdefault("PROTOCOL_BUFFERS_PYTHON_IMPLEMENTATION", "python")

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import torch
import torch.nn.functional as F
from openpyxl import load_workbook
from torch import nn

PROJECT_ROOT = Path(__file__).resolve().parents[3]

from isa.device_models.flash_transistor import ekv_current
from isa.operators.cim import CIMLinear
from isa.vision.config import ModelConfig, apply_preset
from isa.vision.data import get_dataloaders
from isa.vision.models import build_model

PERCENTILES = (1, 5, 10, 50, 90, 95, 99)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Measured-IV verify-write deployment for all ViT fc1 layers"
    )
    parser.add_argument("--model", choices=("physical_vit", "hybrid_vit"), default="hybrid_vit")
    parser.add_argument("--scale", default="small")
    parser.add_argument("--checkpoint", default="")
    parser.add_argument("--data_dir", default=str(PROJECT_ROOT / "data" / "cifar100"))
    parser.add_argument("--iv_excel", default=str(PROJECT_ROOT / "measured_iv.xlsx"))
    parser.add_argument("--iv_sheet", default="cell1")
    parser.add_argument(
        "--assignment_sheet",
        default="",
        help="Reference sheet used to design curve-pair assignments; empty means --iv_sheet",
    )
    parser.add_argument(
        "--sheet_transfer_mode",
        choices=("same_sheet", "same_state", "single_point_remap", "external_map"),
        default="same_sheet",
        help=(
            "same_sheet designs/deploys on one sheet; same_state transfers Value ids; "
            "single_point_remap maps reference states to deployment states at V_write; "
            "external_map loads a precomputed noise-aware global-codebook mapping"
        ),
    )
    parser.add_argument(
        "--state_map_json",
        default="",
        help="JSON mapping produced by build_global_codebooks.py for external_map mode",
    )
    parser.add_argument(
        "--measured_current_scale",
        type=float,
        default=1e-9,
        help="Scale applied to spreadsheet currents; FG50 sheets are nA, so default is 1e-9",
    )
    parser.add_argument(
        "--min_valid_current_na",
        type=float,
        default=None,
        help=(
            "Optional raw-sheet current threshold in nA-like units. Non-finite and "
            "smaller values are left-censored to 0 A in the inference LUT."
        ),
    )
    parser.add_argument(
        "--output_root",
        default=str(PROJECT_ROOT / "measured_iv_deployment_results"),
    )
    parser.add_argument("--experiment_name", default="")
    parser.add_argument("--deploy_layers", default="all", help="all, last, or comma-separated block ids")
    parser.add_argument(
        "--assignment_mode",
        choices=(
            "independent_branch",
            "pairwise_diff",
            "distribution_pairwise",
            "random",
            "histogram_shuffle",
        ),
        default="pairwise_diff",
    )
    parser.add_argument(
        "--shuffle_base_mode",
        choices=("independent_branch", "pairwise_diff"),
        default="pairwise_diff",
    )
    parser.add_argument("--lambda_cm", type=float, default=0.01)
    parser.add_argument(
        "--distribution_points",
        type=int,
        default=9,
        help="Equal-probability Vin quadrature points for distribution_pairwise",
    )
    parser.add_argument(
        "--write_verify_voltage",
        default="auto",
        help="Single hardware write/verify voltage in V, or auto for best robust state spacing",
    )
    parser.add_argument(
        "--vverify_strategy",
        choices=("median", "min_assignment_error"),
        default="min_assignment_error",
    )
    parser.add_argument("--vverify_sample_devices", type=int, default=200000)
    parser.add_argument("--calib_batches", type=int, default=20)
    parser.add_argument("--calib_batch_size", type=int, default=128)
    parser.add_argument("--calib_hist_bins", type=int, default=2000)
    parser.add_argument("--eval_batch_size", type=int, default=128)
    parser.add_argument("--max_eval_batches", type=int, default=0)
    parser.add_argument("--num_workers", type=int, default=2)
    parser.add_argument("--assignment_chunk_devices", type=int, default=8192)
    parser.add_argument("--seed", type=int, default=42)
    parser.add_argument(
        "--inference_backend",
        choices=("torch_segmented_gemm",),
        default="torch_segmented_gemm",
    )
    parser.add_argument(
        "--vin_lut_bins",
        type=int,
        default=0,
        help="Optional approximate Vin LUT; 0 keeps exact measured-grid segmented GEMM",
    )
    parser.add_argument("--torch_compile", action="store_true")
    parser.add_argument("--correctness_rows", type=int, default=4)
    parser.add_argument("--correctness_outputs", type=int, default=64)
    parser.add_argument("--plot_samples_per_layer", type=int, default=5000)
    return parser.parse_args()


def default_checkpoint(model_name: str) -> Path:
    stem = "small_physical_vit/physical_vit" if model_name == "physical_vit" else "small_hybrid_vit/hybrid_vit"
    return PROJECT_ROOT / "outputs" / "cifar100_300ep" / stem / "best_checkpoint.pth"


def parse_deploy_layers(spec: str, depth: int) -> list[int]:
    text = spec.strip().lower()
    if text == "all":
        return list(range(depth))
    if text == "last":
        return [depth - 1]
    try:
        result = sorted({int(part.strip()) for part in text.split(",") if part.strip()})
    except ValueError as exc:
        raise ValueError(f"Invalid --deploy_layers={spec!r}") from exc
    if not result or result[0] < 0 or result[-1] >= depth:
        raise ValueError(f"deploy layer ids must be within [0, {depth - 1}], got {result}")
    return result


def lambda_tag(value: float) -> str:
    return f"{value:g}".replace("-", "m").replace(".", "p")


def experiment_dir_name(args: argparse.Namespace, deploy_ids: Sequence[int], depth: int) -> str:
    if args.experiment_name:
        return args.experiment_name
    if list(deploy_ids) == list(range(depth)):
        layer_tag = "all"
    elif list(deploy_ids) == [depth - 1]:
        layer_tag = "last"
    else:
        layer_tag = "layers_" + "_".join(str(x) for x in deploy_ids)
    name = f"{args.iv_sheet}_{layer_tag}_fc1_{args.assignment_mode}_layerwise_vverify"
    if args.assignment_mode in {"pairwise_diff", "distribution_pairwise", "histogram_shuffle"}:
        name += f"_lambda{lambda_tag(args.lambda_cm)}"
    if args.assignment_mode == "distribution_pairwise":
        name += f"_q{args.distribution_points}_singlewrite{args.write_verify_voltage.replace('.', 'p')}"
    if args.vin_lut_bins:
        name += f"_vinbins{args.vin_lut_bins}"
    if args.min_valid_current_na is not None:
        name += f"_min{lambda_tag(args.min_valid_current_na)}nA"
    return name


def write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    if not rows:
        return
    keys: list[str] = []
    for row in rows:
        for key in row:
            if key not in keys:
                keys.append(key)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=keys)
        writer.writeheader()
        writer.writerows(rows)


class IVLibrary:
    def __init__(
        self,
        vg: torch.Tensor,
        currents_a: torch.Tensor,
        value_names: Sequence[str],
        current_scale: float,
        raw_min: float,
        raw_max: float,
        min_valid_current_raw: float | None,
        raw_point_count: int,
        censored_point_count: int,
        negative_point_count: int,
        nonfinite_or_missing_point_count: int,
    ) -> None:
        self.vg = vg.float().contiguous()
        self.currents_a = currents_a.float().contiguous()
        self.value_names = list(value_names)
        self.current_scale = float(current_scale)
        self.raw_min = float(raw_min)
        self.raw_max = float(raw_max)
        self.min_valid_current_raw = (
            None if min_valid_current_raw is None else float(min_valid_current_raw)
        )
        self.raw_point_count = int(raw_point_count)
        self.censored_point_count = int(censored_point_count)
        self.negative_point_count = int(negative_point_count)
        self.nonfinite_or_missing_point_count = int(nonfinite_or_missing_point_count)
        positive = self.currents_a[self.currents_a > 0]
        if positive.numel() == 0:
            raise ValueError("Measured IV library contains no positive current")
        self.min_positive_a = float(positive.min().item())
        # This is deliberately computed only after conversion to ampere.
        self.i_floor_a = max(1e-12, 0.1 * self.min_positive_a)
        if self.vg.numel() < 2 or not bool(torch.all(self.vg[1:] > self.vg[:-1])):
            raise ValueError("Measured Vg grid must be strictly increasing")
        if self.currents_a.shape != (len(self.value_names), self.vg.numel()):
            raise ValueError("Measured current table shape does not match headers/Vg grid")

    @classmethod
    def load(
        cls,
        path: Path,
        sheet: str,
        current_scale: float,
        min_valid_current_raw: float | None = None,
    ) -> IVLibrary:
        workbook = load_workbook(path, read_only=True, data_only=True)
        if sheet not in workbook.sheetnames:
            raise KeyError(f"Sheet {sheet!r} not found; choices: {workbook.sheetnames}")
        rows = list(workbook[sheet].iter_rows(values_only=True))
        if not rows:
            raise ValueError(f"Sheet {sheet!r} is empty")
        headers = list(rows[0])
        curve_columns = [
            index for index, value in enumerate(headers) if str(value).startswith("Value_")
        ]
        if not curve_columns:
            raise ValueError("No Value_* measured-curve columns found")
        value_names = [str(headers[index]) for index in curve_columns]
        vg_values: list[float] = []
        current_rows: list[list[float]] = []
        missing_count = 0
        for row in rows[1:]:
            if row[0] is None:
                continue
            vg_values.append(float(row[0]))
            current_row: list[float] = []
            for index in curve_columns:
                value = row[index] if index < len(row) else None
                if value is None:
                    current_row.append(float("nan"))
                    missing_count += 1
                else:
                    try:
                        current_row.append(float(value))
                    except (TypeError, ValueError):
                        current_row.append(float("nan"))
                        missing_count += 1
            current_rows.append(current_row)
        raw = torch.tensor(current_rows, dtype=torch.float64).t().contiguous()
        finite = torch.isfinite(raw)
        finite_raw = raw[finite]
        if finite_raw.numel() == 0:
            raise ValueError(f"Sheet {sheet!r} contains no finite measured currents")
        negative_count = int(((raw < 0) & finite).sum().item())
        clean_raw = raw.clone()
        clean_raw[~finite] = 0.0
        if min_valid_current_raw is None:
            censored = ~finite
        else:
            censored = (~finite) | (raw < float(min_valid_current_raw))
            clean_raw[censored] = 0.0
        return cls(
            torch.tensor(vg_values, dtype=torch.float64),
            (clean_raw * float(current_scale)).float(),
            value_names,
            current_scale,
            float(finite_raw.min().item()),
            float(finite_raw.max().item()),
            min_valid_current_raw,
            int(raw.numel()),
            int(censored.sum().item()),
            negative_count,
            max(missing_count, int((~finite).sum().item())),
        )

    def cleaning_row(self, sheet: str, role: str) -> dict[str, Any]:
        return {
            "sheet": sheet,
            "role": role,
            "raw_point_count": self.raw_point_count,
            "min_valid_current_raw_nA_like": self.min_valid_current_raw,
            "censored_point_count": self.censored_point_count,
            "censored_point_ratio": self.censored_point_count / max(self.raw_point_count, 1),
            "negative_point_count": self.negative_point_count,
            "nonfinite_or_missing_point_count": self.nonfinite_or_missing_point_count,
            "raw_current_min_nA_like": self.raw_min,
            "raw_current_max_nA_like": self.raw_max,
            "clean_min_positive_current_A": self.min_positive_a,
            "log_comparison_I_floor_A": self.i_floor_a,
        }

    @property
    def n_curves(self) -> int:
        return int(self.currents_a.shape[0])

    @property
    def n_segments(self) -> int:
        return int(self.vg.numel() - 1)

    def currents_at(self, voltage: float, device: torch.device) -> torch.Tensor:
        vg = self.vg.to(device)
        currents = self.currents_a.to(device)
        value = torch.tensor(float(voltage), dtype=torch.float32, device=device).clamp(vg[0], vg[-1])
        upper = torch.searchsorted(vg, value, right=True).clamp(1, vg.numel() - 1)
        lower = upper - 1
        alpha = (value - vg[lower]) / (vg[upper] - vg[lower])
        return currents[:, lower] + alpha * (currents[:, upper] - currents[:, lower])


def select_single_write_verify_voltage(
    library: IVLibrary, specification: str
) -> tuple[float, list[dict[str, Any]], list[dict[str, Any]]]:
    """Choose a single write voltage where all Value states are positive and ordered.

    Auto mode maximizes the fifth-percentile adjacent-state log-current gap.
    The full scan is returned for auditability; it is independent of the
    distribution-weighted assignment voltage points.
    """
    scan_rows: list[dict[str, Any]] = []
    eligible: list[dict[str, Any]] = []
    expected_descending = torch.arange(library.n_curves - 1, -1, -1)
    for grid_id, voltage_tensor in enumerate(library.vg):
        voltage = float(voltage_tensor.item())
        currents = library.currents_a[:, grid_id]
        all_positive = bool(torch.all(currents > 0))
        sorted_ids = torch.argsort(currents)
        monotonic = bool(torch.equal(sorted_ids.cpu(), expected_descending))
        log_currents = torch.log10(currents.clamp_min(library.i_floor_a))
        adjacent_gaps = torch.abs(log_currents[:-1] - log_currents[1:])
        row = {
            "voltage_V": voltage,
            "all_states_positive": all_positive,
            "monotonic_Value_id_order": monotonic,
            "positive_state_count": int((currents > 0).sum().item()),
            "min_adjacent_log_gap": float(adjacent_gaps.min().item()),
            "p5_adjacent_log_gap": float(
                torch.quantile(adjacent_gaps, torch.tensor(0.05)).item()
            ),
            "median_adjacent_log_gap": float(adjacent_gaps.median().item()),
            "log_current_dynamic_range": float(log_currents.max().item() - log_currents.min().item()),
        }
        scan_rows.append(row)
        if all_positive and monotonic:
            eligible.append(row)
    if specification.strip().lower() == "auto":
        if not eligible:
            raise RuntimeError("No measured Vg point has all-positive, monotonic Value states")
        selected = max(eligible, key=lambda row: row["p5_adjacent_log_gap"])
        voltage = float(selected["voltage_V"])
    else:
        voltage = float(specification)
        if voltage < float(library.vg[0]) or voltage > float(library.vg[-1]):
            raise ValueError(f"write verify voltage {voltage} is outside measured Vg range")
    selected_currents = library.currents_at(voltage, torch.device("cpu"))
    if not bool(torch.all(selected_currents > 0)):
        warnings.warn(
            f"Some measured states are non-positive at write verify voltage {voltage:.3f} V"
        )
    selected_logs = torch.log10(selected_currents.clamp_min(library.i_floor_a))
    state_rows: list[dict[str, Any]] = []
    for value_id, value_name in enumerate(library.value_names):
        gap = (
            float(torch.abs(selected_logs[value_id] - selected_logs[value_id + 1]).item())
            if value_id + 1 < library.n_curves
            else None
        )
        state_rows.append(
            {
                "Value_id": value_id,
                "Value_name": value_name,
                "write_verify_voltage_V": voltage,
                "target_current_A": float(selected_currents[value_id].item()),
                "target_current_nA": float(selected_currents[value_id].item() * 1e9),
                "log10_current_A": float(selected_logs[value_id].item()),
                "log_gap_to_next_Value_id": gap,
            }
        )
    return voltage, scan_rows, state_rows


def build_state_transfer_map(
    assignment_library: IVLibrary,
    deployment_library: IVLibrary,
    voltage: float,
    mode: str,
) -> tuple[torch.Tensor, list[dict[str, Any]], dict[str, Any]]:
    if assignment_library.n_curves != deployment_library.n_curves:
        raise ValueError("Assignment and deployment sheets have different state counts")
    source = assignment_library.currents_at(voltage, torch.device("cpu"))
    target = deployment_library.currents_at(voltage, torch.device("cpu"))
    source_log = torch.log10(source.clamp_min(assignment_library.i_floor_a))
    target_log = torch.log10(target.clamp_min(deployment_library.i_floor_a))
    if mode in {"same_sheet", "same_state"}:
        mapping = torch.arange(assignment_library.n_curves, dtype=torch.long)
    elif mode == "single_point_remap":
        mapping = torch.abs(source_log[:, None] - target_log[None, :]).argmin(dim=1)
    else:
        raise ValueError(f"Unsupported sheet transfer mode {mode}")

    selected = target[mapping]
    selected_log = target_log[mapping]
    log_error = torch.abs(source_log - selected_log)
    rows: list[dict[str, Any]] = []
    for source_id in range(assignment_library.n_curves):
        target_id = int(mapping[source_id].item())
        rows.append(
            {
                "source_Value_id": source_id,
                "source_Value_name": assignment_library.value_names[source_id],
                "deployment_Value_id": target_id,
                "deployment_Value_name": deployment_library.value_names[target_id],
                "write_verify_voltage_V": float(voltage),
                "source_target_current_A": float(source[source_id].item()),
                "deployment_selected_current_A": float(selected[source_id].item()),
                "absolute_log10_current_error": float(log_error[source_id].item()),
                "Value_id_shift": target_id - source_id,
                "state_changed": target_id != source_id,
            }
        )
    summary = {
        "write_verify_voltage_V": float(voltage),
        "state_count": assignment_library.n_curves,
        "changed_state_count": int((mapping != torch.arange(mapping.numel())).sum().item()),
        "changed_state_ratio": float(
            (mapping != torch.arange(mapping.numel())).float().mean().item()
        ),
        "mean_absolute_Value_id_shift": float(
            torch.abs(mapping - torch.arange(mapping.numel())).float().mean().item()
        ),
        "max_absolute_Value_id_shift": int(
            torch.abs(mapping - torch.arange(mapping.numel())).max().item()
        ),
        "mean_write_verify_log10_error": float(log_error.mean().item()),
        "max_write_verify_log10_error": float(log_error.max().item()),
    }
    return mapping, rows, summary


@torch.inference_mode()
def apply_state_transfer_to_layer(
    block_id: int,
    assignment: dict[str, Any],
    state_map: torch.Tensor,
    assignment_library: IVLibrary,
    deployment_library: IVLibrary,
    distribution_points: Sequence[float] | None,
    distribution_weights: Sequence[float] | None,
) -> dict[str, Any]:
    source_pos = assignment["pos_idx"].long()
    source_neg = assignment["neg_idx"].long()
    deployed_pos = state_map[source_pos]
    deployed_neg = state_map[source_neg]
    assignment["deployment_pos_idx"] = deployed_pos.to(torch.int16)
    assignment["deployment_neg_idx"] = deployed_neg.to(torch.int16)
    assignment["deployment_pos_hist"] = torch.bincount(
        deployed_pos.flatten(), minlength=deployment_library.n_curves
    ).tolist()
    assignment["deployment_neg_hist"] = torch.bincount(
        deployed_neg.flatten(), minlength=deployment_library.n_curves
    ).tolist()
    row: dict[str, Any] = {
        "layer_name": f"blocks.{block_id}.mlp.fc1",
        "number_of_diff_pairs": int(source_pos.numel()),
        "source_pos_state_changed_ratio": float((deployed_pos != source_pos).float().mean().item()),
        "source_neg_state_changed_ratio": float((deployed_neg != source_neg).float().mean().item()),
        "used_deployment_Value_id_count": int(
            torch.unique(torch.cat((deployed_pos.flatten(), deployed_neg.flatten()))).numel()
        ),
    }
    if distribution_points and distribution_weights:
        source_grid = torch.stack(
            [
                assignment_library.currents_at(float(point), torch.device("cpu"))
                for point in distribution_points
            ],
            dim=0,
        )
        deployment_grid = torch.stack(
            [
                deployment_library.currents_at(float(point), torch.device("cpu"))
                for point in distribution_points
            ],
            dim=0,
        )
        weights = torch.tensor(distribution_weights, dtype=torch.float32)
        weights = weights / weights.sum()
        source_diff = source_grid[:, source_pos.flatten()] - source_grid[:, source_neg.flatten()]
        deployment_diff = (
            deployment_grid[:, deployed_pos.flatten()] - deployment_grid[:, deployed_neg.flatten()]
        )
        diff_change_by_point = (source_diff - deployment_diff).abs().mean(dim=1)
        source_abs_by_point = source_diff.abs().mean(dim=1)
        weighted_change = float((weights * diff_change_by_point).sum().item())
        weighted_source = float((weights * source_abs_by_point).sum().item())
        row.update(
            {
                "transfer_distribution_mean_abs_diff_change_A": weighted_change,
                "transfer_distribution_relative_diff_change": weighted_change
                / max(weighted_source, assignment_library.i_floor_a),
            }
        )
    return row


class VoltageAccumulator:
    def __init__(self, voltage_max: float, hist_bins: int) -> None:
        self.voltage_max = float(voltage_max)
        self.hist_bins = int(hist_bins)
        self.count = 0
        self.total = 0.0
        self.total_sq = 0.0
        self.minimum = math.inf
        self.maximum = -math.inf
        self.raw_total = 0.0
        self.raw_total_sq = 0.0
        self.raw_minimum = math.inf
        self.raw_maximum = -math.inf
        self.clamp_low = 0
        self.clamp_high = 0
        self.hist = torch.zeros(self.hist_bins, dtype=torch.float64)
        self.sample_vin: torch.Tensor | None = None
        self.mapping_to_fc1_verified = True
        self._last_output_ptr: int | None = None
        self._last_output_shape: tuple[int, ...] | None = None

    @torch.no_grad()
    def update(self, module: nn.Module, module_input: tuple[torch.Tensor, ...], output: torch.Tensor) -> None:
        source = module_input[0].detach().float()
        raw = source * module.scale.detach().float() + module.shift.detach().float()
        values = output.detach().float()
        count = values.numel()
        self.count += count
        self.total += float(values.double().sum().item())
        self.total_sq += float(values.double().square().sum().item())
        self.minimum = min(self.minimum, float(values.min().item()))
        self.maximum = max(self.maximum, float(values.max().item()))
        self.raw_total += float(raw.double().sum().item())
        self.raw_total_sq += float(raw.double().square().sum().item())
        self.raw_minimum = min(self.raw_minimum, float(raw.min().item()))
        self.raw_maximum = max(self.raw_maximum, float(raw.max().item()))
        self.clamp_low += int((raw < 0.0).sum().item())
        self.clamp_high += int((raw > self.voltage_max).sum().item())
        self.hist += torch.histc(
            values,
            bins=self.hist_bins,
            min=0.0,
            max=self.voltage_max,
        ).double().cpu()
        if self.sample_vin is None:
            self.sample_vin = values[:1, :4].cpu().contiguous()
        self._last_output_ptr = output.data_ptr()
        self._last_output_shape = tuple(output.shape)

    @torch.no_grad()
    def verify_fc1_input(
        self, _module: nn.Module, module_input: tuple[torch.Tensor, ...]
    ) -> None:
        value = module_input[0]
        current = value.data_ptr() == self._last_output_ptr and tuple(value.shape) == self._last_output_shape
        self.mapping_to_fc1_verified = self.mapping_to_fc1_verified and bool(current)

    def quantile(self, probability: float) -> float:
        if self.count == 0:
            return float("nan")
        cumulative = torch.cumsum(self.hist, dim=0)
        target = probability * max(float(cumulative[-1].item()), 1.0)
        index = int(torch.searchsorted(cumulative, torch.tensor(target, dtype=cumulative.dtype)).item())
        index = min(max(index, 0), self.hist_bins - 1)
        return self.voltage_max * (index + 0.5) / self.hist_bins

    def density(self, voltage: float) -> float:
        index = int(float(voltage) / self.voltage_max * self.hist_bins)
        index = min(max(index, 0), self.hist_bins - 1)
        lo = max(0, index - 2)
        hi = min(self.hist_bins, index + 3)
        return float(self.hist[lo:hi].sum().item()) / max(float(self.hist.sum().item()), 1.0)

    def row(self, layer_name: str) -> dict[str, Any]:
        count = max(self.count, 1)
        mean = self.total / count
        raw_mean = self.raw_total / count
        variance = max(self.total_sq / count - mean * mean, 0.0)
        raw_variance = max(self.raw_total_sq / count - raw_mean * raw_mean, 0.0)
        row: dict[str, Any] = {
            "layer_name": layer_name,
            "vin_min": self.minimum,
            "vin_max": self.maximum,
            "vin_mean": mean,
            "vin_std": math.sqrt(variance),
            "vin_unclamped_min": self.raw_minimum,
            "vin_unclamped_max": self.raw_maximum,
            "vin_unclamped_mean": raw_mean,
            "vin_unclamped_std": math.sqrt(raw_variance),
            "clamp_low_ratio": self.clamp_low / count,
            "clamp_high_ratio": self.clamp_high / count,
            "clamp_ratio": (self.clamp_low + self.clamp_high) / count,
            "sample_count": self.count,
            "fc1_input_is_mapping_output": self.mapping_to_fc1_verified,
        }
        for percentile in PERCENTILES:
            row[f"vin_p{percentile}"] = self.quantile(percentile / 100.0)
        return row


def distribution_quadrature(
    accumulator: VoltageAccumulator, point_count: int
) -> tuple[list[float], list[float], list[str]]:
    """Approximate the empirical Vin distribution with equal-probability points."""
    if point_count < 2:
        raise ValueError("--distribution_points must be at least 2")
    buckets: dict[float, dict[str, Any]] = {}
    for point_id in range(point_count):
        probability = (point_id + 0.5) / point_count
        voltage = round(float(accumulator.quantile(probability)), 6)
        bucket = buckets.setdefault(voltage, {"weight": 0.0, "quantiles": []})
        bucket["weight"] += 1.0 / point_count
        bucket["quantiles"].append(probability)
    points = list(buckets.keys())
    weights = [float(buckets[point]["weight"]) for point in points]
    labels = [
        ",".join(f"{probability:.6f}" for probability in buckets[point]["quantiles"])
        for point in points
    ]
    return points, weights, labels


@torch.inference_mode()
def calibrate_voltages(
    model: nn.Module,
    loader: Iterable[tuple[torch.Tensor, torch.Tensor]],
    batches: int,
    hist_bins: int,
    device: torch.device,
) -> tuple[list[dict[str, Any]], dict[int, VoltageAccumulator]]:
    accumulators: dict[int, VoltageAccumulator] = {}
    handles = []
    for block_id, block in enumerate(model.blocks):
        accumulator = VoltageAccumulator(block.mlp.input_mapping.voltage_max, hist_bins)
        accumulators[block_id] = accumulator
        handles.append(block.mlp.input_mapping.register_forward_hook(accumulator.update))
        handles.append(block.mlp.fc1.register_forward_pre_hook(accumulator.verify_fc1_input))
    try:
        for batch_id, (images, _) in enumerate(loader):
            if batch_id >= batches:
                break
            model(images.to(device, non_blocking=True))
            print(f"  calibration batch {batch_id + 1}/{batches}", flush=True)
    finally:
        for handle in handles:
            handle.remove()
    rows = [
        accumulators[index].row(f"blocks.{index}.mlp.fc1") for index in range(len(model.blocks))
    ]
    if not all(row["fc1_input_is_mapping_output"] for row in rows):
        raise RuntimeError("fc1 input hook did not observe the exact VoltageMapping output")
    return rows, accumulators


@torch.inference_mode()
def select_vverify_for_layer(
    layer_name: str,
    fc1: CIMLinear,
    stats: dict[str, Any],
    accumulator: VoltageAccumulator,
    library: IVLibrary,
    strategy: str,
    sample_devices: int,
    seed: int,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    device = fc1.vth_pos.device
    all_vth = torch.cat((fc1.vth_pos.detach().flatten(), fc1.vth_neg.detach().flatten())).float()
    if all_vth.numel() > sample_devices:
        generator = torch.Generator(device=device)
        generator.manual_seed(seed)
        indices = torch.randperm(all_vth.numel(), generator=generator, device=device)[:sample_devices]
        sampled_vth = all_vth[indices]
    else:
        sampled_vth = all_vth

    if strategy == "median":
        candidates = [float(stats["vin_p50"])]
        candidate_window = "median_interpolated"
    else:
        lower = max(float(stats["vin_p10"]), float(library.vg[0].item()))
        upper = min(float(stats["vin_p90"]), float(library.vg[-1].item()))
        mask = (library.vg >= lower) & (library.vg <= upper)
        candidate_tensor = library.vg[mask]
        candidate_window = "p10_p90"
        if candidate_tensor.numel() < 3:
            lower = max(float(stats["vin_p5"]), float(library.vg[0].item()))
            upper = min(float(stats["vin_p95"]), float(library.vg[-1].item()))
            candidate_tensor = library.vg[(library.vg >= lower) & (library.vg <= upper)]
            candidate_window = "p5_p95"
        if candidate_tensor.numel() == 0:
            nearest = int(torch.argmin(torch.abs(library.vg - float(stats["vin_p50"]))).item())
            candidate_tensor = library.vg[nearest : nearest + 1]
            candidate_window = "nearest_to_median"
        candidates = [float(value) for value in candidate_tensor.tolist()]

    rows: list[dict[str, Any]] = []
    for candidate in candidates:
        measured = library.currents_at(candidate, device)
        measured_log = torch.log10(measured.clamp_min(library.i_floor_a))
        target = ekv_current(
            torch.full_like(sampled_vth, candidate), sampled_vth, fc1.cfg
        ).float()
        target_log = torch.log10(target.clamp_min(library.i_floor_a))
        nearest_error = torch.abs(target_log[:, None] - measured_log[None, :]).amin(dim=1)
        measured_min = float(measured_log.min().item())
        measured_max = float(measured_log.max().item())
        out_of_range = ((target_log < measured_min) | (target_log > measured_max)).float().mean()
        density = accumulator.density(candidate)
        score = float(nearest_error.mean().item()) + 5.0 * float(out_of_range.item())
        if strategy == "min_assignment_error":
            score -= 0.01 * math.log(density + 1e-6)
        rows.append(
            {
                "layer_name": layer_name,
                "v_candidate": candidate,
                "candidate_window": candidate_window,
                "score": score,
                "mean_nearest_log_error": float(nearest_error.mean().item()),
                "out_of_range_ratio": float(out_of_range.item()),
                "input_density": density,
                "measured_logI_min": measured_min,
                "measured_logI_max": measured_max,
                "target_logI_min": float(target_log.min().item()),
                "target_logI_max": float(target_log.max().item()),
            }
        )
    selected = min(rows, key=lambda row: row["score"])
    selection = dict(selected)
    selection.update(
        {
            "V_verify": selection.pop("v_candidate"),
            "strategy": strategy,
            "vin_p10": stats["vin_p10"],
            "vin_p50": stats["vin_p50"],
            "vin_p90": stats["vin_p90"],
            "sampled_branches": int(sampled_vth.numel()),
        }
    )
    return selection, rows


@torch.inference_mode()
def select_all_vverify(
    model: nn.Module,
    stats_rows: Sequence[dict[str, Any]],
    accumulators: dict[int, VoltageAccumulator],
    library: IVLibrary,
    args: argparse.Namespace,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    selections: list[dict[str, Any]] = []
    candidates: list[dict[str, Any]] = []
    for block_id, block in enumerate(model.blocks):
        selection, candidate_rows = select_vverify_for_layer(
            f"blocks.{block_id}.mlp.fc1",
            block.mlp.fc1,
            stats_rows[block_id],
            accumulators[block_id],
            library,
            args.vverify_strategy,
            args.vverify_sample_devices,
            args.seed + block_id,
        )
        selections.append(selection)
        candidates.extend(candidate_rows)
        print(
            f"  block {block_id:2d}: V_verify={selection['V_verify']:.4f} V "
            f"score={selection['score']:.5f} p10/p50/p90="
            f"{selection['vin_p10']:.3f}/{selection['vin_p50']:.3f}/{selection['vin_p90']:.3f}",
            flush=True,
        )
    return selections, candidates


def _compute_assignment_indices(
    target_pos: torch.Tensor,
    target_neg: torch.Tensor,
    measured: torch.Tensor,
    i_floor: float,
    mode: str,
    lambda_cm: float,
    chunk_size: int,
    seed: int,
) -> tuple[torch.Tensor, torch.Tensor]:
    device = target_pos.device
    flat_pos = target_pos.flatten()
    flat_neg = target_neg.flatten()
    n_pairs = flat_pos.numel()
    n_curves = measured.numel()
    if mode == "random":
        generator = torch.Generator(device=device)
        generator.manual_seed(seed)
        return (
            torch.randint(n_curves, (n_pairs,), generator=generator, device=device),
            torch.randint(n_curves, (n_pairs,), generator=generator, device=device),
        )

    if mode == "histogram_shuffle":
        raise ValueError("histogram_shuffle must be applied after its base assignment")

    pos_indices = torch.empty(n_pairs, dtype=torch.long, device=device)
    neg_indices = torch.empty_like(pos_indices)
    if mode == "independent_branch":
        measured_log = torch.log10(measured.clamp_min(i_floor))
        for start in range(0, n_pairs, chunk_size):
            end = min(start + chunk_size, n_pairs)
            pos_log = torch.log10(flat_pos[start:end].clamp_min(i_floor))
            neg_log = torch.log10(flat_neg[start:end].clamp_min(i_floor))
            pos_indices[start:end] = torch.abs(pos_log[:, None] - measured_log[None, :]).argmin(dim=1)
            neg_indices[start:end] = torch.abs(neg_log[:, None] - measured_log[None, :]).argmin(dim=1)
        return pos_indices, neg_indices

    if mode != "pairwise_diff":
        raise ValueError(f"Unsupported assignment mode {mode}")
    pos_pair_ids = torch.arange(n_curves, device=device).repeat_interleave(n_curves)
    neg_pair_ids = torch.arange(n_curves, device=device).repeat(n_curves)
    scale = max(
        float(measured.abs().max().item()),
        float(flat_pos.abs().max().item()),
        float(flat_neg.abs().max().item()),
        i_floor,
    )
    pair_diff = (measured[pos_pair_ids] - measured[neg_pair_ids]) / scale
    pair_sum = (measured[pos_pair_ids] + measured[neg_pair_ids]) / scale
    for start in range(0, n_pairs, chunk_size):
        end = min(start + chunk_size, n_pairs)
        target_diff = (flat_pos[start:end] - flat_neg[start:end]) / scale
        target_sum = (flat_pos[start:end] + flat_neg[start:end]) / scale
        loss = (target_diff[:, None] - pair_diff[None, :]).square()
        if lambda_cm != 0.0:
            loss.add_((target_sum[:, None] - pair_sum[None, :]).square(), alpha=lambda_cm)
        best = loss.argmin(dim=1)
        pos_indices[start:end] = pos_pair_ids[best]
        neg_indices[start:end] = neg_pair_ids[best]
    return pos_indices, neg_indices


def _compute_distribution_pairwise_indices(
    target_pos_grid: torch.Tensor,
    target_neg_grid: torch.Tensor,
    measured_grid: torch.Tensor,
    weights: torch.Tensor,
    i_floor: float,
    lambda_cm: float,
    chunk_size: int,
) -> tuple[torch.Tensor, torch.Tensor]:
    """Match full differential-current shapes using weighted squared distance.

    Shapes are [Q, N] for target branches and [Q, K] for measured curves.
    Weighted-distance expansion reduces each device-to-codebook comparison to
    two small-Q GEMMs instead of materializing [chunk, Q, K*K].
    """
    device = target_pos_grid.device
    q_points, n_pairs = target_pos_grid.shape
    if target_neg_grid.shape != (q_points, n_pairs):
        raise ValueError("target branch grids have inconsistent shapes")
    if measured_grid.shape[0] != q_points or weights.numel() != q_points:
        raise ValueError("distribution grid, measured grid, and weights do not align")
    n_curves = measured_grid.shape[1]
    pos_pair_ids = torch.arange(n_curves, device=device).repeat_interleave(n_curves)
    neg_pair_ids = torch.arange(n_curves, device=device).repeat(n_curves)
    scale = max(
        float(measured_grid.abs().max().item()),
        float(target_pos_grid.abs().max().item()),
        float(target_neg_grid.abs().max().item()),
        i_floor,
    )
    sqrt_weights = torch.sqrt(weights.float() / weights.sum()).view(1, q_points)
    pair_diff = (
        (measured_grid[:, pos_pair_ids] - measured_grid[:, neg_pair_ids]).t() / scale
    )
    pair_sum = (
        (measured_grid[:, pos_pair_ids] + measured_grid[:, neg_pair_ids]).t() / scale
    )
    pair_diff_w = pair_diff * sqrt_weights
    pair_sum_w = pair_sum * sqrt_weights
    pair_diff_norm = pair_diff_w.square().sum(dim=1)
    pair_sum_norm = pair_sum_w.square().sum(dim=1)
    target_diff = ((target_pos_grid - target_neg_grid).t() / scale).contiguous()
    target_sum = ((target_pos_grid + target_neg_grid).t() / scale).contiguous()
    pos_indices = torch.empty(n_pairs, dtype=torch.long, device=device)
    neg_indices = torch.empty_like(pos_indices)
    for start in range(0, n_pairs, chunk_size):
        end = min(start + chunk_size, n_pairs)
        target_diff_w = target_diff[start:end] * sqrt_weights
        loss = pair_diff_norm.unsqueeze(0) - 2.0 * (target_diff_w @ pair_diff_w.t())
        if lambda_cm != 0.0:
            target_sum_w = target_sum[start:end] * sqrt_weights
            common_loss = pair_sum_norm.unsqueeze(0) - 2.0 * (target_sum_w @ pair_sum_w.t())
            loss.add_(common_loss, alpha=lambda_cm)
        best = loss.argmin(dim=1)
        pos_indices[start:end] = pos_pair_ids[best]
        neg_indices[start:end] = neg_pair_ids[best]
    return pos_indices, neg_indices


@torch.inference_mode()
def assign_layer(
    block_id: int,
    fc1: CIMLinear,
    v_verify: float,
    library: IVLibrary,
    args: argparse.Namespace,
    distribution_points: Sequence[float] | None = None,
    distribution_weights: Sequence[float] | None = None,
) -> dict[str, Any]:
    device = fc1.vth_pos.device
    shape = tuple(fc1.vth_pos.shape)
    gate = torch.tensor(float(v_verify), device=device, dtype=torch.float32)
    target_pos = ekv_current(gate, fc1.vth_pos.detach().float(), fc1.cfg)
    target_neg = ekv_current(gate, fc1.vth_neg.detach().float(), fc1.cfg)
    measured = library.currents_at(v_verify, device)
    mode = args.assignment_mode
    distribution_metrics: dict[str, Any] = {}
    if mode == "distribution_pairwise":
        if not distribution_points or not distribution_weights:
            raise ValueError("distribution_pairwise requires Vin points and weights")
        points_tensor = torch.tensor(
            list(distribution_points), device=device, dtype=torch.float32
        )
        weights_tensor = torch.tensor(
            list(distribution_weights), device=device, dtype=torch.float32
        )
        weights_tensor = weights_tensor / weights_tensor.sum()
        flat_pos_vth = fc1.vth_pos.detach().float().flatten()
        flat_neg_vth = fc1.vth_neg.detach().float().flatten()
        target_pos_grid = ekv_current(
            points_tensor[:, None], flat_pos_vth[None, :], fc1.cfg
        )
        target_neg_grid = ekv_current(
            points_tensor[:, None], flat_neg_vth[None, :], fc1.cfg
        )
        measured_grid = torch.stack(
            [library.currents_at(float(point), device) for point in distribution_points],
            dim=0,
        )
        pos_idx, neg_idx = _compute_distribution_pairwise_indices(
            target_pos_grid,
            target_neg_grid,
            measured_grid,
            weights_tensor,
            library.i_floor_a,
            args.lambda_cm,
            args.assignment_chunk_devices,
        )
        selected_pos_grid = measured_grid[:, pos_idx]
        selected_neg_grid = measured_grid[:, neg_idx]
        target_diff_grid = target_pos_grid - target_neg_grid
        selected_diff_grid = selected_pos_grid - selected_neg_grid
        target_sum_grid = target_pos_grid + target_neg_grid
        selected_sum_grid = selected_pos_grid + selected_neg_grid
        diff_abs_by_point = (target_diff_grid - selected_diff_grid).abs().mean(dim=1)
        target_abs_by_point = target_diff_grid.abs().mean(dim=1)
        selected_abs_by_point = selected_diff_grid.abs().mean(dim=1)
        common_abs_by_point = (target_sum_grid - selected_sum_grid).abs().mean(dim=1)
        target_common_by_point = target_sum_grid.abs().mean(dim=1)
        distribution_diff_error = float((weights_tensor * diff_abs_by_point).sum().item())
        distribution_target_abs = float((weights_tensor * target_abs_by_point).sum().item())
        distribution_common_error = float((weights_tensor * common_abs_by_point).sum().item())
        distribution_target_common = float((weights_tensor * target_common_by_point).sum().item())
        distribution_metrics = {
            "assignment_voltage_points_V": json.dumps(list(distribution_points)),
            "assignment_voltage_weights": json.dumps(list(distribution_weights)),
            "distribution_mean_abs_target_diff_current_A": distribution_target_abs,
            "distribution_mean_abs_selected_diff_current_A": float(
                (weights_tensor * selected_abs_by_point).sum().item()
            ),
            "distribution_mean_abs_diff_error_A": distribution_diff_error,
            "distribution_relative_diff_error": distribution_diff_error
            / max(distribution_target_abs, library.i_floor_a),
            "distribution_rms_diff_error_A": float(
                torch.sqrt(
                    (
                        weights_tensor
                        * (target_diff_grid - selected_diff_grid).square().mean(dim=1)
                    ).sum()
                ).item()
            ),
            "distribution_mean_abs_common_mode_error_A": distribution_common_error,
            "distribution_relative_common_mode_error": distribution_common_error
            / max(distribution_target_common, library.i_floor_a),
            "distribution_point_relative_diff_errors": json.dumps(
                [
                    float(diff_abs_by_point[index].item())
                    / max(float(target_abs_by_point[index].item()), library.i_floor_a)
                    for index in range(points_tensor.numel())
                ]
            ),
        }
        del (
            target_pos_grid,
            target_neg_grid,
            measured_grid,
            selected_pos_grid,
            selected_neg_grid,
            target_diff_grid,
            selected_diff_grid,
            target_sum_grid,
            selected_sum_grid,
        )
    else:
        base_mode = args.shuffle_base_mode if mode == "histogram_shuffle" else mode
        pos_idx, neg_idx = _compute_assignment_indices(
            target_pos,
            target_neg,
            measured,
            library.i_floor_a,
            base_mode,
            args.lambda_cm,
            args.assignment_chunk_devices,
            args.seed + 1000 * block_id,
        )
    if mode == "histogram_shuffle":
        generator = torch.Generator(device=device)
        generator.manual_seed(args.seed + 1000 * block_id + 17)
        pos_idx = pos_idx[torch.randperm(pos_idx.numel(), generator=generator, device=device)]
        neg_idx = neg_idx[torch.randperm(neg_idx.numel(), generator=generator, device=device)]

    flat_target_pos = target_pos.flatten()
    flat_target_neg = target_neg.flatten()
    selected_pos = measured[pos_idx]
    selected_neg = measured[neg_idx]
    target_pos_log = torch.log10(flat_target_pos.clamp_min(library.i_floor_a))
    target_neg_log = torch.log10(flat_target_neg.clamp_min(library.i_floor_a))
    selected_pos_log = torch.log10(selected_pos.clamp_min(library.i_floor_a))
    selected_neg_log = torch.log10(selected_neg.clamp_min(library.i_floor_a))
    target_log = torch.cat((target_pos_log, target_neg_log))
    selected_log = torch.cat((selected_pos_log, selected_neg_log))
    branch_error = torch.abs(target_log - selected_log)
    target_diff = flat_target_pos - flat_target_neg
    selected_diff = selected_pos - selected_neg
    diff_error = torch.abs(target_diff - selected_diff)
    same = pos_idx == neg_idx
    mean_abs_target = float(target_diff.abs().mean().item())
    relative_diff_error = float(diff_error.mean().item()) / max(mean_abs_target, library.i_floor_a)
    same_threshold = max(library.i_floor_a, 0.01 * mean_abs_target)
    same_near_zero = (
        float((target_diff[same].abs() <= same_threshold).float().mean().item())
        if bool(same.any())
        else float("nan")
    )
    measured_log = torch.log10(measured.clamp_min(library.i_floor_a))
    out_of_range = (
        (target_log < measured_log.min()) | (target_log > measured_log.max())
    ).float().mean()
    union_ids = torch.unique(torch.cat((pos_idx, neg_idx)))
    layer_stats: dict[str, Any] = {
        "layer_name": f"blocks.{block_id}.mlp.fc1",
        "V_verify": float(v_verify),
        "number_of_diff_pairs": int(target_diff.numel()),
        "target_logI_min": float(target_log.min().item()),
        "target_logI_max": float(target_log.max().item()),
        "target_logI_mean": float(target_log.mean().item()),
        "target_logI_std": float(target_log.std(unbiased=False).item()),
        "selected_logI_min": float(selected_log.min().item()),
        "selected_logI_max": float(selected_log.max().item()),
        "selected_logI_mean": float(selected_log.mean().item()),
        "selected_logI_std": float(selected_log.std(unbiased=False).item()),
        "mean_branch_log_error": float(branch_error.mean().item()),
        "max_branch_log_error": float(branch_error.max().item()),
        "mean_abs_target_diff_current_A": mean_abs_target,
        "mean_abs_selected_diff_current_A": float(selected_diff.abs().mean().item()),
        "mean_abs_diff_error_A": float(diff_error.mean().item()),
        "relative_diff_error": relative_diff_error,
        "ratio_same_curve": float(same.float().mean().item()),
        "same_curve_target_near_zero_ratio": same_near_zero,
        "same_curve_near_zero_threshold_A": same_threshold,
        "used_Value_id_count": int(union_ids.numel()),
        "out_of_range_ratio": float(out_of_range.item()),
        "measured_logI_min": float(measured_log.min().item()),
        "measured_logI_max": float(measured_log.max().item()),
    }
    layer_stats.update(distribution_metrics)
    pos_hist = torch.bincount(pos_idx, minlength=library.n_curves).cpu().tolist()
    neg_hist = torch.bincount(neg_idx, minlength=library.n_curves).cpu().tolist()
    delta_hist = torch.bincount(
        pos_idx - neg_idx + library.n_curves - 1,
        minlength=2 * library.n_curves - 1,
    ).cpu().tolist()

    sample_count = min(args.plot_samples_per_layer, target_diff.numel())
    sample_gen = torch.Generator(device=device)
    sample_gen.manual_seed(args.seed + 2000 * block_id)
    sample_ids = torch.randperm(target_diff.numel(), generator=sample_gen, device=device)[:sample_count]
    plot_sample = {
        "target_log": torch.cat((target_pos_log[sample_ids], target_neg_log[sample_ids])).cpu().numpy(),
        "selected_log": torch.cat((selected_pos_log[sample_ids], selected_neg_log[sample_ids])).cpu().numpy(),
        "target_diff": target_diff[sample_ids].cpu().numpy(),
        "selected_diff": selected_diff[sample_ids].cpu().numpy(),
    }
    return {
        "pos_idx": pos_idx.reshape(shape).to(torch.int16).cpu(),
        "neg_idx": neg_idx.reshape(shape).to(torch.int16).cpu(),
        "stats": layer_stats,
        "pos_hist": pos_hist,
        "neg_hist": neg_hist,
        "delta_hist": delta_hist,
        "plot_sample": plot_sample,
    }


class MeasuredIVSegmentedLinear(nn.Module):
    """Exact measured-curve interpolation using LUT coefficients and segmented GEMMs."""

    def __init__(
        self,
        original: CIMLinear,
        pos_idx_cpu: torch.Tensor,
        neg_idx_cpu: torch.Tensor,
        library: IVLibrary,
        vin_lut_bins: int = 0,
    ) -> None:
        super().__init__()
        device = original.vth_pos.device
        self.in_features = int(original.in_features)
        self.out_features = int(original.out_features)
        self.vmin = float(original.cfg.get("V_signed_min", -4.0))
        self.vmax = float(original.cfg.get("V_signed_max", 4.0))
        lookup_min = max(float(library.vg[0].item()), float(original.cfg.get("V_min", 0.0)))
        lookup_max = min(float(library.vg[-1].item()), float(original.cfg.get("V_max", 4.0)))
        self.lookup_min = lookup_min
        self.lookup_max = lookup_max
        self.vin_lut_bins = int(vin_lut_bins)
        self.register_buffer("r_tia_value", original.r_tia.detach().float().clone())
        self.register_buffer("pos_idx", pos_idx_cpu.to(device=device, dtype=torch.int16))
        self.register_buffer("neg_idx", neg_idx_cpu.to(device=device, dtype=torch.int16))
        pos_idx = self.pos_idx.long()
        neg_idx = self.neg_idx.long()
        vg = library.vg.to(device)
        currents = library.currents_a.to(device)
        self.register_buffer("vg", vg)
        self.register_buffer("measured_currents", currents)

        if self.vin_lut_bins > 0:
            centers = torch.linspace(lookup_min, lookup_max, self.vin_lut_bins, device=device)
            diff_current = torch.empty(
                self.vin_lut_bins,
                self.out_features,
                self.in_features,
                dtype=torch.float32,
                device=device,
            )
            for bin_id, center in enumerate(centers.tolist()):
                measured = library.currents_at(center, device)
                diff_current[bin_id] = measured[pos_idx] - measured[neg_idx]
            self.register_buffer("bin_centers", centers)
            self.register_buffer("diff_current", diff_current)
            self.register_buffer("active_segments", torch.empty(0, dtype=torch.long, device=device))
            self.register_buffer("diff_a", torch.empty(0, device=device))
            self.register_buffer("diff_b", torch.empty(0, device=device))
        else:
            dv = vg[1:] - vg[:-1]
            slopes = (currents[:, 1:] - currents[:, :-1]) / dv.unsqueeze(0)
            intercepts = currents[:, :-1] - slopes * vg[:-1].unsqueeze(0)
            active = torch.nonzero(
                (vg[:-1] <= lookup_max) & (vg[1:] >= lookup_min), as_tuple=False
            ).flatten()
            diff_a = torch.empty(
                active.numel(),
                self.out_features,
                self.in_features,
                dtype=torch.float32,
                device=device,
            )
            diff_b = torch.empty_like(diff_a)
            for local_id, segment_id in enumerate(active.tolist()):
                diff_a[local_id] = slopes[:, segment_id][pos_idx] - slopes[:, segment_id][neg_idx]
                diff_b[local_id] = (
                    intercepts[:, segment_id][pos_idx] - intercepts[:, segment_id][neg_idx]
                )
            self.register_buffer("active_segments", active)
            self.register_buffer("diff_a", diff_a)
            self.register_buffer("diff_b", diff_b)
            self.register_buffer("bin_centers", torch.empty(0, device=device))
            self.register_buffer("diff_current", torch.empty(0, device=device))

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        original_dtype = x.dtype
        shape = x.shape
        flat = x.float().reshape(-1, self.in_features).clamp(self.lookup_min, self.lookup_max)
        output = torch.zeros(
            flat.shape[0], self.out_features, dtype=torch.float32, device=flat.device
        )
        if self.vin_lut_bins > 0:
            if self.vin_lut_bins == 1:
                bin_id = torch.zeros_like(flat, dtype=torch.long)
            else:
                scale = (self.vin_lut_bins - 1) / (self.lookup_max - self.lookup_min)
                bin_id = torch.round((flat - self.lookup_min) * scale).long().clamp(
                    0, self.vin_lut_bins - 1
                )
            for current_bin in range(self.vin_lut_bins):
                mask = (bin_id == current_bin).to(flat.dtype)
                output.add_(F.linear(mask, self.diff_current[current_bin]))
        else:
            segment_id = torch.searchsorted(self.vg, flat, right=True).sub_(1).clamp_(
                0, self.vg.numel() - 2
            )
            for local_id in range(self.active_segments.numel()):
                current_segment = self.active_segments[local_id]
                mask = (segment_id == current_segment).to(flat.dtype)
                output.add_(F.linear(flat * mask, self.diff_a[local_id]))
                output.add_(F.linear(mask, self.diff_b[local_id]))
        output.mul_(self.r_tia_value.float()).clamp_(self.vmin, self.vmax)
        return output.reshape(*shape[:-1], self.out_features).to(original_dtype)


class CompileFallback(nn.Module):
    """Use torch.compile when available, reverting to eager on any compile/runtime failure."""

    def __init__(self, eager: nn.Module) -> None:
        super().__init__()
        self.eager = eager
        self.compile_active = False
        compiled = None
        if hasattr(torch, "compile"):
            try:
                compiled = torch.compile(eager, dynamic=False)
                self.compile_active = True
            except Exception as exc:  # noqa: BLE001  # pragma: no cover - version dependent
                warnings.warn(f"torch.compile setup failed; using eager: {exc}")
        object.__setattr__(self, "_compiled", compiled)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        compiled = object.__getattribute__(self, "_compiled")
        if compiled is not None:
            try:
                return compiled(x)
            except Exception as exc:  # noqa: BLE001  # pragma: no cover - version dependent
                warnings.warn(f"torch.compile execution failed; using eager: {exc}")
                object.__setattr__(self, "_compiled", None)
                self.compile_active = False
        return self.eager(x)


@torch.inference_mode()
def slow_reference(
    layer: MeasuredIVSegmentedLinear,
    x: torch.Tensor,
    output_count: int,
) -> torch.Tensor:
    flat = x.float().reshape(-1, layer.in_features).clamp(layer.lookup_min, layer.lookup_max)
    if layer.vin_lut_bins > 0:
        if layer.vin_lut_bins == 1:
            quantized = torch.full_like(flat, float(layer.bin_centers[0].item()))
        else:
            scale = (layer.vin_lut_bins - 1) / (layer.lookup_max - layer.lookup_min)
            ids = torch.round((flat - layer.lookup_min) * scale).long().clamp(
                0, layer.vin_lut_bins - 1
            )
            quantized = layer.bin_centers[ids]
        flat = quantized
    upper = torch.searchsorted(layer.vg, flat, right=True).clamp(1, layer.vg.numel() - 1)
    lower = upper - 1
    alpha = (flat - layer.vg[lower]) / (layer.vg[upper] - layer.vg[lower])
    flat_lower = lower.reshape(-1)
    flat_upper = upper.reshape(-1)
    values_low = layer.measured_currents[:, flat_lower].t().reshape(
        flat.shape[0], layer.in_features, -1
    )
    values_high = layer.measured_currents[:, flat_upper].t().reshape_as(values_low)
    all_values = values_low + alpha.unsqueeze(-1) * (values_high - values_low)
    pos = layer.pos_idx[:output_count].long()
    neg = layer.neg_idx[:output_count].long()
    source = all_values.unsqueeze(1).expand(-1, output_count, -1, -1)
    pos_values = torch.gather(
        source, 3, pos.unsqueeze(0).unsqueeze(-1).expand(flat.shape[0], -1, -1, -1)
    ).squeeze(-1)
    neg_values = torch.gather(
        source, 3, neg.unsqueeze(0).unsqueeze(-1).expand(flat.shape[0], -1, -1, -1)
    ).squeeze(-1)
    output = (pos_values - neg_values).sum(dim=2) * layer.r_tia_value.float()
    return output.clamp(layer.vmin, layer.vmax)


@torch.inference_mode()
def check_layer_correctness(
    block_id: int,
    layer: MeasuredIVSegmentedLinear,
    sample_vin: torch.Tensor,
    rows: int,
    output_count: int,
    device: torch.device,
) -> dict[str, Any]:
    sample = sample_vin.reshape(-1, layer.in_features)[:rows].reshape(1, -1, layer.in_features).to(device)
    count = min(output_count, layer.out_features)
    fast = layer(sample).reshape(-1, layer.out_features)[:, :count]
    reference = slow_reference(layer, sample, count)
    difference = (fast - reference).abs()
    relative = float(torch.linalg.vector_norm(fast - reference).item()) / max(
        float(torch.linalg.vector_norm(reference).item()), 1e-12
    )
    row = {
        "layer_name": f"blocks.{block_id}.mlp.fc1",
        "max_abs_error": float(difference.max().item()),
        "mean_abs_error": float(difference.mean().item()),
        "relative_error": relative,
        "rows_checked": int(fast.shape[0]),
        "outputs_checked": count,
    }
    if not math.isfinite(row["max_abs_error"]) or row["max_abs_error"] > 1e-5 or relative > 1e-4:
        raise RuntimeError(f"Measured-IV backend correctness check failed: {row}")
    return row


@torch.inference_mode()
def evaluate_model(
    model: nn.Module,
    loader: Iterable[tuple[torch.Tensor, torch.Tensor]],
    device: torch.device,
    max_batches: int,
) -> dict[str, Any]:
    model.eval()
    if device.type == "cuda":
        torch.cuda.empty_cache()
        torch.cuda.reset_peak_memory_stats(device)
        torch.cuda.synchronize(device)
    started = time.perf_counter()
    correct = 0
    total = 0
    batches = 0
    for batch_id, (images, targets) in enumerate(loader):
        if max_batches > 0 and batch_id >= max_batches:
            break
        logits = model(images.to(device, non_blocking=True))
        targets = targets.to(device, non_blocking=True)
        correct += int((logits.argmax(dim=1) == targets).sum().item())
        total += int(targets.numel())
        batches += 1
    if device.type == "cuda":
        torch.cuda.synchronize(device)
    elapsed = time.perf_counter() - started
    peak = int(torch.cuda.max_memory_allocated(device)) if device.type == "cuda" else None
    return {
        "accuracy": 100.0 * correct / max(total, 1),
        "correct": correct,
        "number_of_images": total,
        "number_of_batches": batches,
        "total_inference_time_s": elapsed,
        "images_per_sec": total / max(elapsed, 1e-12),
        "peak_memory_bytes": peak,
    }


def save_plots(
    output_dir: Path,
    assignments: dict[int, dict[str, Any]],
    library: IVLibrary,
) -> None:
    layer_ids = sorted(assignments)
    overall_pos = np.sum([assignments[index]["pos_hist"] for index in layer_ids], axis=0)
    overall_neg = np.sum([assignments[index]["neg_hist"] for index in layer_ids], axis=0)
    heatmap = np.asarray(
        [
            np.asarray(assignments[index]["pos_hist"]) + np.asarray(assignments[index]["neg_hist"])
            for index in layer_ids
        ],
        dtype=np.float64,
    )
    heatmap /= np.maximum(heatmap.sum(axis=1, keepdims=True), 1.0)
    fig, axes = plt.subplots(2, 1, figsize=(12, 8), constrained_layout=True)
    x = np.arange(library.n_curves)
    axes[0].bar(x, overall_pos, label="pos", alpha=0.8)
    axes[0].bar(x, overall_neg, bottom=overall_pos, label="neg", alpha=0.8)
    axes[0].set_xlabel("Value_id")
    axes[0].set_ylabel("selected branches")
    axes[0].legend()
    image = axes[1].imshow(heatmap, aspect="auto", interpolation="nearest", cmap="viridis")
    axes[1].set_yticks(np.arange(len(layer_ids)), [str(index) for index in layer_ids])
    axes[1].set_xlabel("Value_id")
    axes[1].set_ylabel("block")
    fig.colorbar(image, ax=axes[1], label="within-layer ratio")
    fig.savefig(output_dir / "selected_value_histogram.png", dpi=180)
    plt.close(fig)

    target_log = np.concatenate([assignments[index]["plot_sample"]["target_log"] for index in layer_ids])
    selected_log = np.concatenate([assignments[index]["plot_sample"]["selected_log"] for index in layer_ids])
    fig, ax = plt.subplots(figsize=(7, 7), constrained_layout=True)
    ax.scatter(target_log, selected_log, s=2, alpha=0.15, rasterized=True)
    lo = min(float(target_log.min()), float(selected_log.min()))
    hi = max(float(target_log.max()), float(selected_log.max()))
    ax.plot([lo, hi], [lo, hi], "r--", linewidth=1)
    ax.set_xlabel("target log10(I/A)")
    ax.set_ylabel("selected log10(I/A)")
    ax.set_title("Target vs selected measured branch current")
    fig.savefig(output_dir / "target_vs_selected_logI.png", dpi=180)
    plt.close(fig)

    target_diff = np.concatenate([assignments[index]["plot_sample"]["target_diff"] for index in layer_ids])
    selected_diff = np.concatenate([assignments[index]["plot_sample"]["selected_diff"] for index in layer_ids])
    fig, axes = plt.subplots(1, 2, figsize=(13, 5), constrained_layout=True)
    axes[0].scatter(target_diff * 1e9, selected_diff * 1e9, s=2, alpha=0.15, rasterized=True)
    axes[0].set_xlabel("target differential current (nA)")
    axes[0].set_ylabel("selected differential current (nA)")
    axes[0].set_title("Differential current target vs selected")
    abs_target = np.log10(np.maximum(np.abs(target_diff), library.i_floor_a))
    abs_selected = np.log10(np.maximum(np.abs(selected_diff), library.i_floor_a))
    axes[1].hist(abs_target, bins=80, alpha=0.55, density=True, label="target")
    axes[1].hist(abs_selected, bins=80, alpha=0.55, density=True, label="selected")
    axes[1].set_xlabel("log10(max(abs(I_diff), I_floor) / A)")
    axes[1].set_ylabel("density")
    axes[1].legend()
    fig.savefig(output_dir / "diff_current_target_vs_selected.png", dpi=180)
    plt.close(fig)


def weighted_mean(rows: Sequence[dict[str, Any]], key: str, weight_key: str) -> float:
    numerator = sum(float(row[key]) * float(row[weight_key]) for row in rows)
    denominator = sum(float(row[weight_key]) for row in rows)
    return numerator / max(denominator, 1.0)


def main() -> None:
    args = parse_args()
    random.seed(args.seed)
    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    if torch.cuda.is_available():
        torch.cuda.manual_seed_all(args.seed)
    torch.backends.cuda.matmul.allow_tf32 = False
    torch.backends.cudnn.allow_tf32 = False
    torch.set_float32_matmul_precision("highest")
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    cfg = ModelConfig()
    apply_preset(cfg, args.scale)
    deploy_ids = parse_deploy_layers(args.deploy_layers, cfg.depth)
    output_dir = Path(args.output_root) / experiment_dir_name(args, deploy_ids, cfg.depth)
    output_dir.mkdir(parents=True, exist_ok=True)
    checkpoint_path = Path(args.checkpoint) if args.checkpoint else default_checkpoint(args.model)
    print(f"device={device} output={output_dir}", flush=True)
    print(f"checkpoint={checkpoint_path}", flush=True)

    assignment_sheet = args.assignment_sheet or args.iv_sheet
    if args.sheet_transfer_mode == "same_sheet" and assignment_sheet != args.iv_sheet:
        raise ValueError(
            "--sheet_transfer_mode=same_sheet requires --assignment_sheet to equal --iv_sheet"
        )
    assignment_library = IVLibrary.load(
        Path(args.iv_excel),
        assignment_sheet,
        args.measured_current_scale,
        args.min_valid_current_na,
    )
    deployment_library = IVLibrary.load(
        Path(args.iv_excel),
        args.iv_sheet,
        args.measured_current_scale,
        args.min_valid_current_na,
    )
    library = assignment_library
    cleaning_rows = [assignment_library.cleaning_row(assignment_sheet, "assignment")]
    if args.iv_sheet != assignment_sheet:
        cleaning_rows.append(deployment_library.cleaning_row(args.iv_sheet, "deployment"))
    write_csv(output_dir / "iv_cleaning_stats.csv", cleaning_rows)
    print(
        f"assignment IV={assignment_sheet}: {library.n_curves} curves, "
        f"{library.vg.numel()} Vg points, raw=[{library.raw_min:g},{library.raw_max:g}] "
        f"nA-like units, censored={library.censored_point_count}/{library.raw_point_count}, "
        f"scale={library.current_scale:g}, I_floor_A={library.i_floor_a:.6g}",
        flush=True,
    )
    print(
        f"deployment IV={args.iv_sheet}: raw=[{deployment_library.raw_min:g},"
        f"{deployment_library.raw_max:g}] nA-like units, "
        f"censored={deployment_library.censored_point_count}/"
        f"{deployment_library.raw_point_count}, I_floor_A={deployment_library.i_floor_a:.6g}",
        flush=True,
    )
    write_verify_voltage: float | None = None
    if args.sheet_transfer_mode != "external_map" and (
        args.assignment_mode == "distribution_pairwise"
        or args.sheet_transfer_mode == "single_point_remap"
    ):
        write_verify_voltage, write_scan_rows, write_state_rows = (
            select_single_write_verify_voltage(library, args.write_verify_voltage)
        )
        write_csv(output_dir / "write_verify_voltage_scan.csv", write_scan_rows)
        write_csv(output_dir / "write_verify_state_targets.csv", write_state_rows)
        selected_scan = min(
            write_scan_rows,
            key=lambda row: abs(float(row["voltage_V"]) - write_verify_voltage),
        )
        print(
            f"single-point hardware write verify: V_write={write_verify_voltage:.3f} V, "
            f"all_positive={selected_scan['all_states_positive']}, "
            f"monotonic={selected_scan['monotonic_Value_id_order']}, "
            f"p5_adjacent_log_gap={selected_scan['p5_adjacent_log_gap']:.6f}",
            flush=True,
        )
    if args.sheet_transfer_mode == "external_map":
        if not args.state_map_json:
            raise ValueError("--state_map_json is required for external_map mode")
        state_map_payload = json.loads(Path(args.state_map_json).read_text(encoding="utf-8"))
        mapping_values = state_map_payload.get("mapping")
        if not isinstance(mapping_values, list):
            raise ValueError("external state-map JSON must contain a list named 'mapping'")
        state_map = torch.tensor(mapping_values, dtype=torch.long)
        if state_map.numel() != assignment_library.n_curves:
            raise ValueError(
                f"external map has {state_map.numel()} states, expected "
                f"{assignment_library.n_curves}"
            )
        if int(state_map.min()) < 0 or int(state_map.max()) >= deployment_library.n_curves:
            raise ValueError("external map contains deployment Value ids outside the sheet")
        state_transfer_rows = state_map_payload.get("rows", [])
        state_transfer_summary = dict(state_map_payload.get("summary", {}))
        state_transfer_summary.setdefault("write_verify_voltage_V", None)
        state_transfer_summary.setdefault("state_count", assignment_library.n_curves)
        state_transfer_summary.setdefault("mapping_method", "external_map")
        write_csv(output_dir / "state_transfer_map.csv", state_transfer_rows)
        print(
            f"sheet transfer=external_map: {assignment_sheet} -> {args.iv_sheet}, "
            f"states={state_map.numel()}, target_states={deployment_library.n_curves}, "
            f"method={state_transfer_summary.get('mapping_method')}",
            flush=True,
        )
    elif write_verify_voltage is None:
        state_map = torch.arange(library.n_curves, dtype=torch.long)
        state_transfer_rows: list[dict[str, Any]] = []
        state_transfer_summary: dict[str, Any] = {
            "write_verify_voltage_V": None,
            "state_count": library.n_curves,
            "changed_state_count": 0,
            "changed_state_ratio": 0.0,
            "mean_absolute_Value_id_shift": 0.0,
            "max_absolute_Value_id_shift": 0,
            "mean_write_verify_log10_error": None,
            "max_write_verify_log10_error": None,
        }
    else:
        state_map, state_transfer_rows, state_transfer_summary = build_state_transfer_map(
            assignment_library,
            deployment_library,
            write_verify_voltage,
            args.sheet_transfer_mode,
        )
        write_csv(output_dir / "state_transfer_map.csv", state_transfer_rows)
        print(
            f"sheet transfer={args.sheet_transfer_mode}: {assignment_sheet} -> {args.iv_sheet}, "
            f"changed_states={state_transfer_summary['changed_state_count']}/"
            f"{state_transfer_summary['state_count']}, "
            f"mean_write_log_error="
            f"{state_transfer_summary['mean_write_verify_log10_error']:.6f}",
            flush=True,
        )
    checkpoint = torch.load(checkpoint_path, map_location="cpu")
    state_dict = checkpoint.get("model", checkpoint)
    model = build_model(args.model, cfg)
    model.load_state_dict(state_dict, strict=True)
    model.to(device).eval()
    for block_id, block in enumerate(model.blocks):
        if not isinstance(block.mlp.fc1, CIMLinear):
            raise TypeError(f"blocks.{block_id}.mlp.fc1 is not CIMLinear")
        if tuple(block.mlp.fc1.vth_pos.shape) != (1536, 384) and args.scale == "small":
            raise ValueError(f"Unexpected small fc1 shape {tuple(block.mlp.fc1.vth_pos.shape)}")
    fc2_type = type(model.blocks[0].mlp.fc2).__name__
    print(f"loaded checkpoint strictly; fc2 type is {fc2_type}", flush=True)

    _, calibration_loader = get_dataloaders(
        args.data_dir,
        batch_size=args.calib_batch_size,
        num_workers=args.num_workers,
        distributed=False,
    )
    print("Calibrating actual VoltageMapping -> fc1 inputs...", flush=True)
    input_stats, accumulators = calibrate_voltages(
        model,
        calibration_loader,
        args.calib_batches,
        args.calib_hist_bins,
        device,
    )
    write_csv(output_dir / "layerwise_input_voltage_stats.csv", input_stats)
    distribution_points_by_layer: dict[int, list[float]] = {}
    distribution_weights_by_layer: dict[int, list[float]] = {}
    if args.assignment_mode == "distribution_pairwise":
        quadrature_rows: list[dict[str, Any]] = []
        for block_id in range(cfg.depth):
            points, weights, quantile_labels = distribution_quadrature(
                accumulators[block_id], args.distribution_points
            )
            distribution_points_by_layer[block_id] = points
            distribution_weights_by_layer[block_id] = weights
            for point_id, (point, weight, label) in enumerate(
                zip(points, weights, quantile_labels)
            ):
                quadrature_rows.append(
                    {
                        "layer_name": f"blocks.{block_id}.mlp.fc1",
                        "point_id": point_id,
                        "vin_V": point,
                        "probability_weight": weight,
                        "source_quantile_centers": label,
                    }
                )
            print(
                f"  block {block_id:2d} distribution points: "
                + ", ".join(f"{point:.3f}V@{weight:.3f}" for point, weight in zip(points, weights)),
                flush=True,
            )
        write_csv(output_dir / "distribution_assignment_points.csv", quadrature_rows)

    print("Selecting per-layer V_verify...", flush=True)
    vverify_rows, candidate_rows = select_all_vverify(
        model, input_stats, accumulators, library, args
    )
    write_csv(output_dir / "layerwise_vverify_selection.csv", vverify_rows)
    write_csv(output_dir / "vverify_candidate_scores.csv", candidate_rows)
    vverify_by_layer = {
        row["layer_name"]: float(row["V_verify"]) for row in vverify_rows
    }

    _, evaluation_loader = get_dataloaders(
        args.data_dir,
        batch_size=args.eval_batch_size,
        num_workers=args.num_workers,
        distributed=False,
    )
    print("Evaluating original continuous checkpoint...", flush=True)
    continuous_eval = evaluate_model(model, evaluation_loader, device, args.max_eval_batches)
    print(
        f"continuous accuracy={continuous_eval['accuracy']:.4f}% "
        f"speed={continuous_eval['images_per_sec']:.2f} images/s",
        flush=True,
    )

    assignments: dict[int, dict[str, Any]] = {}
    assignment_rows: list[dict[str, Any]] = []
    transfer_layer_rows: list[dict[str, Any]] = []
    histogram_rows: list[dict[str, Any]] = []
    deployment_histogram_rows: list[dict[str, Any]] = []
    delta_histogram: Counter[int] = Counter()
    for block_id in deploy_ids:
        started = time.perf_counter()
        result = assign_layer(
            block_id,
            model.blocks[block_id].mlp.fc1,
            vverify_rows[block_id]["V_verify"],
            library,
            args,
            distribution_points_by_layer.get(block_id),
            distribution_weights_by_layer.get(block_id),
        )
        result["stats"]["assignment_time_s"] = time.perf_counter() - started
        assignments[block_id] = result
        assignment_rows.append(result["stats"])
        transfer_row = apply_state_transfer_to_layer(
            block_id,
            result,
            state_map,
            assignment_library,
            deployment_library,
            distribution_points_by_layer.get(block_id),
            distribution_weights_by_layer.get(block_id),
        )
        transfer_layer_rows.append(transfer_row)
        for value_id in range(library.n_curves):
            for branch, histogram in (("pos", result["pos_hist"]), ("neg", result["neg_hist"])):
                histogram_rows.append(
                    {
                        "layer_name": f"blocks.{block_id}.mlp.fc1",
                        "branch": branch,
                        "Value_id": value_id,
                        "count": int(histogram[value_id]),
                        "ratio": int(histogram[value_id]) / result["stats"]["number_of_diff_pairs"],
                    }
                )
        for value_id in range(deployment_library.n_curves):
            for branch, histogram in (
                ("pos", result["deployment_pos_hist"]),
                ("neg", result["deployment_neg_hist"]),
            ):
                deployment_histogram_rows.append(
                    {
                        "layer_name": f"blocks.{block_id}.mlp.fc1",
                        "branch": branch,
                        "Value_id": value_id,
                        "count": int(histogram[value_id]),
                        "ratio": int(histogram[value_id])
                        / result["stats"]["number_of_diff_pairs"],
                    }
                )
        for offset, count in enumerate(result["delta_hist"]):
            delta_histogram[offset - library.n_curves + 1] += int(count)
        distribution_message = ""
        if "distribution_relative_diff_error" in result["stats"]:
            distribution_message = (
                f" distribution_rel_diff_err="
                f"{result['stats']['distribution_relative_diff_error']:.5f}"
            )
        print(
            f"  assigned block {block_id:2d}: same={result['stats']['ratio_same_curve']:.4f} "
            f"branch_log_err={result['stats']['mean_branch_log_error']:.5f} "
            f"reference_point_rel_diff_err={result['stats']['relative_diff_error']:.5f}"
            f"{distribution_message} transfer_rel_diff_change="
            f"{transfer_row.get('transfer_distribution_relative_diff_change', float('nan')):.5f}",
            flush=True,
        )
    write_csv(output_dir / "assignment_stats.csv", histogram_rows)
    write_csv(output_dir / "deployment_assignment_stats.csv", deployment_histogram_rows)
    write_csv(output_dir / "layerwise_assignment_stats.csv", assignment_rows)
    write_csv(output_dir / "layerwise_transfer_stats.csv", transfer_layer_rows)
    write_csv(
        output_dir / "pos_minus_neg_curve_histogram.csv",
        [{"k_pos_minus_k_neg": key, "count": value} for key, value in sorted(delta_histogram.items())],
    )
    save_plots(output_dir, assignments, library)

    print("Building measured-IV segmented-GEMM layers and checking numerics...", flush=True)
    correctness_rows: list[dict[str, Any]] = []
    compile_wrappers: list[CompileFallback] = []
    for block_id in deploy_ids:
        original = model.blocks[block_id].mlp.fc1
        measured_layer = MeasuredIVSegmentedLinear(
            original,
            assignments[block_id]["deployment_pos_idx"],
            assignments[block_id]["deployment_neg_idx"],
            deployment_library,
            args.vin_lut_bins,
        )
        correctness = check_layer_correctness(
            block_id,
            measured_layer,
            accumulators[block_id].sample_vin,
            args.correctness_rows,
            args.correctness_outputs,
            device,
        )
        correctness_rows.append(correctness)
        replacement: nn.Module = measured_layer
        if args.torch_compile:
            wrapper = CompileFallback(measured_layer)
            compile_wrappers.append(wrapper)
            replacement = wrapper
        model.blocks[block_id].mlp.fc1 = replacement
        print(
            f"  block {block_id:2d}: max_abs_error={correctness['max_abs_error']:.3e} "
            f"relative_error={correctness['relative_error']:.3e}",
            flush=True,
        )
    write_csv(output_dir / "correctness_check.csv", correctness_rows)
    if sum(isinstance(model.blocks[index].mlp.fc1, (MeasuredIVSegmentedLinear, CompileFallback)) for index in deploy_ids) != len(deploy_ids):
        raise RuntimeError("Not all requested fc1 layers were replaced")
    for index, block in enumerate(model.blocks):
        if index not in deploy_ids and not isinstance(block.mlp.fc1, CIMLinear):
            raise RuntimeError(f"Non-deployed block {index} was unexpectedly modified")

    print("Evaluating measured-IV deployed model...", flush=True)
    deployed_eval = evaluate_model(model, evaluation_loader, device, args.max_eval_batches)
    print(
        f"deployed accuracy={deployed_eval['accuracy']:.4f}% "
        f"drop={continuous_eval['accuracy'] - deployed_eval['accuracy']:.4f} points "
        f"speed={deployed_eval['images_per_sec']:.2f} images/s",
        flush=True,
    )

    total_pairs = sum(row["number_of_diff_pairs"] for row in assignment_rows)
    overall_hist = np.sum(
        [
            np.asarray(assignments[index]["pos_hist"]) + np.asarray(assignments[index]["neg_hist"])
            for index in deploy_ids
        ],
        axis=0,
    )
    deployment_overall_hist = np.sum(
        [
            np.asarray(assignments[index]["deployment_pos_hist"])
            + np.asarray(assignments[index]["deployment_neg_hist"])
            for index in deploy_ids
        ],
        axis=0,
    )
    layer_histograms = {
        f"blocks.{index}.mlp.fc1": {
            library.value_names[value_id]: int(
                assignments[index]["pos_hist"][value_id]
                + assignments[index]["neg_hist"][value_id]
            )
            for value_id in range(library.n_curves)
        }
        for index in deploy_ids
    }
    deployment_layer_histograms = {
        f"blocks.{index}.mlp.fc1": {
            deployment_library.value_names[value_id]: int(
                assignments[index]["deployment_pos_hist"][value_id]
                + assignments[index]["deployment_neg_hist"][value_id]
            )
            for value_id in range(deployment_library.n_curves)
        }
        for index in deploy_ids
    }
    compile_enabled = bool(args.torch_compile) and bool(compile_wrappers) and all(
        wrapper.compile_active for wrapper in compile_wrappers
    )
    metrics: dict[str, Any] = {
        "checkpoint_path": str(checkpoint_path),
        "checkpoint_epoch": checkpoint.get("epoch"),
        "checkpoint_best_accuracy": checkpoint.get("best_acc"),
        "checkpoint_loaded_strictly": True,
        "model_name": f"{args.model}_{args.scale}",
        "dataset": "CIFAR-100",
        "deploy_layers": deploy_ids,
        "assignment_mode": args.assignment_mode,
        "shuffle_base_mode": args.shuffle_base_mode if args.assignment_mode == "histogram_shuffle" else None,
        "lambda_cm": args.lambda_cm,
        "assignment_objective": (
            "Vin-distribution-weighted multi-point differential/common-mode squared error"
            if args.assignment_mode == "distribution_pairwise"
            else "single-reference-voltage assignment"
        ),
        "distribution_point_count_requested": (
            args.distribution_points if args.assignment_mode == "distribution_pairwise" else None
        ),
        "distribution_assignment_points_by_layer": (
            {
                f"blocks.{index}.mlp.fc1": distribution_points_by_layer[index]
                for index in deploy_ids
            }
            if args.assignment_mode == "distribution_pairwise"
            else None
        ),
        "single_write_verify_voltage_V": write_verify_voltage,
        "single_write_verify_uses_multiple_hardware_reads": False,
        "assignment_iv_sheet": assignment_sheet,
        "deployment_iv_sheet": args.iv_sheet,
        "sheet_transfer_mode": args.sheet_transfer_mode,
        "external_state_map_json": args.state_map_json or None,
        "state_transfer_summary": state_transfer_summary,
        "vverify_strategy": args.vverify_strategy,
        "per_layer_V_verify": vverify_by_layer,
        "iv_excel": str(args.iv_excel),
        "iv_sheet": args.iv_sheet,
        "measured_current_unit_scale": library.current_scale,
        "min_valid_current_raw_nA_like": args.min_valid_current_na,
        "assignment_censored_point_count": assignment_library.censored_point_count,
        "assignment_censored_point_ratio": assignment_library.censored_point_count
        / max(assignment_library.raw_point_count, 1),
        "deployment_censored_point_count": deployment_library.censored_point_count,
        "deployment_censored_point_ratio": deployment_library.censored_point_count
        / max(deployment_library.raw_point_count, 1),
        "measured_raw_current_min": library.raw_min,
        "measured_raw_current_max": library.raw_max,
        "I_floor_A": library.i_floor_a,
        "deployment_measured_raw_current_min": deployment_library.raw_min,
        "deployment_measured_raw_current_max": deployment_library.raw_max,
        "deployment_I_floor_A": deployment_library.i_floor_a,
        "inference_backend": args.inference_backend,
        "vin_lut_bins": args.vin_lut_bins,
        "torch_compile_requested": bool(args.torch_compile),
        "torch_compile_enabled": compile_enabled,
        "number_of_deployed_fc1_layers": len(deploy_ids),
        "number_of_differential_pairs": total_pairs,
        "number_of_physical_branches": 2 * total_pairs,
        "fc2_type_and_preservation": f"{fc2_type}; unchanged from checkpoint",
        "original_continuous_accuracy": continuous_eval["accuracy"],
        "measured_deployed_accuracy": deployed_eval["accuracy"],
        "accuracy_drop": continuous_eval["accuracy"] - deployed_eval["accuracy"],
        "random_assignment_accuracy": deployed_eval["accuracy"] if args.assignment_mode == "random" else None,
        "histogram_preserving_shuffle_accuracy": deployed_eval["accuracy"] if args.assignment_mode == "histogram_shuffle" else None,
        "mean_log10_branch_error": weighted_mean(
            assignment_rows, "mean_branch_log_error", "number_of_diff_pairs"
        ),
        "max_log10_branch_error": max(row["max_branch_log_error"] for row in assignment_rows),
        "mean_absolute_differential_current_error_A": weighted_mean(
            assignment_rows, "mean_abs_diff_error_A", "number_of_diff_pairs"
        ),
        "relative_differential_current_error": weighted_mean(
            assignment_rows, "relative_diff_error", "number_of_diff_pairs"
        ),
        "distribution_relative_differential_current_error": (
            weighted_mean(
                assignment_rows,
                "distribution_relative_diff_error",
                "number_of_diff_pairs",
            )
            if args.assignment_mode == "distribution_pairwise"
            else None
        ),
        "distribution_mean_absolute_differential_current_error_A": (
            weighted_mean(
                assignment_rows,
                "distribution_mean_abs_diff_error_A",
                "number_of_diff_pairs",
            )
            if args.assignment_mode == "distribution_pairwise"
            else None
        ),
        "ratio_same_curve": weighted_mean(
            assignment_rows, "ratio_same_curve", "number_of_diff_pairs"
        ),
        "used_curve_count": int(np.count_nonzero(overall_hist)),
        "deployment_used_curve_count": int(np.count_nonzero(deployment_overall_hist)),
        "selected_Value_id_histogram": {
            library.value_names[index]: int(value) for index, value in enumerate(overall_hist.tolist())
        },
        "layerwise_selected_Value_id_histogram": layer_histograms,
        "deployment_selected_Value_id_histogram": {
            deployment_library.value_names[index]: int(value)
            for index, value in enumerate(deployment_overall_hist.tolist())
        },
        "deployment_layerwise_selected_Value_id_histogram": deployment_layer_histograms,
        "target_logI_range": [
            min(row["target_logI_min"] for row in assignment_rows),
            max(row["target_logI_max"] for row in assignment_rows),
        ],
        "measured_logI_range": [
            min(row["measured_logI_min"] for row in assignment_rows),
            max(row["measured_logI_max"] for row in assignment_rows),
        ],
        "out_of_range_ratio": weighted_mean(
            assignment_rows, "out_of_range_ratio", "number_of_diff_pairs"
        ),
        "transfer_distribution_relative_diff_change": (
            weighted_mean(
                transfer_layer_rows,
                "transfer_distribution_relative_diff_change",
                "number_of_diff_pairs",
            )
            if args.assignment_mode == "distribution_pairwise"
            else None
        ),
        "throughput_images_per_sec": deployed_eval["images_per_sec"],
        "measured_deployment_inference_time_s": deployed_eval["total_inference_time_s"],
        "continuous_EKV_inference_time_s": continuous_eval["total_inference_time_s"],
        "continuous_images_per_sec": continuous_eval["images_per_sec"],
        "evaluated_images": deployed_eval["number_of_images"],
        "correctness_max_abs_error": max(row["max_abs_error"] for row in correctness_rows),
        "correctness_mean_abs_error": float(
            np.mean([row["mean_abs_error"] for row in correctness_rows])
        ),
        "gpu_name": torch.cuda.get_device_name(device) if device.type == "cuda" else None,
    }
    with (output_dir / "metrics.json").open("w", encoding="utf-8") as handle:
        json.dump(metrics, handle, indent=2, ensure_ascii=False)

    speed = {
        "backend_name": args.inference_backend,
        "batch_size": args.eval_batch_size,
        "number_of_images_evaluated": deployed_eval["number_of_images"],
        "total_inference_time_s": deployed_eval["total_inference_time_s"],
        "images_per_sec": deployed_eval["images_per_sec"],
        "gpu_name": metrics["gpu_name"],
        "peak_memory_bytes": deployed_eval["peak_memory_bytes"],
        "continuous_model_total_inference_time_s": continuous_eval["total_inference_time_s"],
        "continuous_model_images_per_sec": continuous_eval["images_per_sec"],
        "continuous_model_peak_memory_bytes": continuous_eval["peak_memory_bytes"],
    }
    with (output_dir / "speed_benchmark.json").open("w", encoding="utf-8") as handle:
        json.dump(speed, handle, indent=2, ensure_ascii=False)

    last_selection = vverify_rows[-1]
    last_warning = ""
    if abs(float(last_selection["V_verify"]) - 1.6) > 0.3:
        last_warning = (
            f"\n警告：blocks.11.fc1 的 V_verify={last_selection['V_verify']:.3f} V，"
            "与旧经验值 1.6 V 偏差较大；已确认 hook 是 VoltageMapping 输出/器件真实输入，"
            "因此优先检查当前 checkpoint 的 Vin 分布与 min_assignment_error 评分。"
        )
    vverify_lines = "\n".join(
        f"  - {row['layer_name']}: {row['V_verify']:.3f} V "
        f"(p10/p50/p90={row['vin_p10']:.3f}/{row['vin_p50']:.3f}/{row['vin_p90']:.3f})"
        for row in vverify_rows
    )
    summary = f"""CIFAR-100 measured-IV 部署总结

1. 关键代码：models/cim_layer.py、models/ekv_function.py、models/physical_vit.py、utils/data.py，以及旧 deploy_*.py 诊断脚本。
2. checkpoint：{checkpoint_path}；使用 model state_dict 严格加载，未训练、未重新初始化（epoch={checkpoint.get('epoch')}, checkpoint best_acc={checkpoint.get('best_acc')}）。
3. fc1 替换：请求并成功替换 {len(deploy_ids)} 层，block ids={deploy_ids}；差分对 {total_pairs:,}，物理分支 {2 * total_pairs:,}。fc2 为 {fc2_type}，保持 checkpoint 原样。
4. 实测单位与清洗：Excel 原始值按 nA × {library.current_scale:g} 转为 A；阈值={args.min_valid_current_na} nA-like，低于阈值、负数及非有限值在推理 LUT 中左删失为 0 A，原始 Excel 不修改。assignment sheet={assignment_sheet}，删失 {assignment_library.censored_point_count}/{assignment_library.raw_point_count} 点；deployment sheet={args.iv_sheet}，删失 {deployment_library.censored_point_count}/{deployment_library.raw_point_count} 点。
5. 每层 V_verify（真实 VoltageMapping clamp 后、fc1 器件输入；指针级 hook 校验通过）：
{vverify_lines}
6. assignment={args.assignment_mode}, lambda_cm={args.lambda_cm:g}；distribution_points={args.distribution_points if args.assignment_mode == 'distribution_pairwise' else None}；sheet transfer={args.sheet_transfer_mode} ({assignment_sheet} -> {args.iv_sheet})；硬件仍为单点 V_write={write_verify_voltage} V。continuous accuracy={continuous_eval['accuracy']:.4f}%，measured accuracy={deployed_eval['accuracy']:.4f}%，drop={continuous_eval['accuracy'] - deployed_eval['accuracy']:.4f} points。
7. random accuracy={metrics['random_assignment_accuracy']}；histogram-preserving shuffle accuracy={metrics['histogram_preserving_shuffle_accuracy']}（非对应模式时为 null，单独实验目录记录）。
8. pos/neg 选择同一曲线比例={metrics['ratio_same_curve']:.6f}；mean branch log10 error={metrics['mean_log10_branch_error']:.6f}；relative differential-current error={metrics['relative_differential_current_error']:.6f}；跨 sheet distribution relative diff change={metrics['transfer_distribution_relative_diff_change']}。
9. 同曲线是否被 pairwise 缓解需与 independent_branch 目录横向比较；本目录逐层统计见 layerwise_assignment_stats.csv，k_pos-k_neg 直方图见 pos_minus_neg_curve_histogram.csv。
10. backend={args.inference_backend}，vin_lut_bins={args.vin_lut_bins}，torch_compile_enabled={compile_enabled}；measured speed={deployed_eval['images_per_sec']:.2f} images/s，continuous speed={continuous_eval['images_per_sec']:.2f} images/s。
11. measured backend 慢速参考校验：max_abs_error={metrics['correctness_max_abs_error']:.3e}，阈值 1e-5，已通过。{last_warning}
"""
    (output_dir / "summary.txt").write_text(summary, encoding="utf-8")
    print(f"All artifacts saved to {output_dir}", flush=True)


if __name__ == "__main__":
    main()
